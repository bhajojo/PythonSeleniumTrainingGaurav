import openpyxl

class readingCell:

    wb = None
    excelFilePath = None

    def __init__(self, excelFilePath):
        self.excelFilePath = excelFilePath
        self.wb = openpyxl.load_workbook(excelFilePath)

    def readColumByIndex(self, sheetNAme, rowIndex, ColIndex):
        sheet = self.wb[sheetNAme]
        return sheet.cell(rowIndex, ColIndex)

    def reafByColName(self, sheetName, rowIndex, colName):
        sheet =self.wb[sheetName]
        colIndex=1;
        while (sheet.cell(1, colIndex ).value() !=""):
            if (colName == sheet.cell(1, colIndex ).value()):
                break
            colIndex = colIndex+1

    def getRowCount(self, sheetName):
        sheet =self.wb[sheetName]
        return sheet.max_row

    def getColCount(self, sheetName):
        sheet =self.wb[sheetName]
        return sheet.max_column

    def writeByColIndex(self, sheetName, rowIndex, colIndex, value, excelFilePath):
        sheet = self.wb[sheetName]
        sheet.cell(rowIndex, colIndex).value = value
        self.wb.save(excelFilePath)




