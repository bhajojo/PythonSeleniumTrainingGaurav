import openpyxl

path = "D:\Training\Python_Selenium_Training\PythonSeleniumTrainingGaurav\Module_9\TestData.xlsx"
# Create Reference to the excel
Workbook = openpyxl.load_workbook(path)
sheet = Workbook.active


maxcol = sheet.max_column

count1=1
while (count1 <= maxcol) :
    print(sheet.cell(1,count1).value)
    count1=count1+1